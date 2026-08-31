from __future__ import annotations

import hashlib
import json
import re
import shutil
from datetime import date, datetime
from zipfile import BadZipFile
from copy import copy
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Callable, Iterable, Iterator

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from .domain import (
    CompiledColumnMapping,
    CompiledMergePlan,
    CompiledOperation,
    CompiledRowMapping,
    CompiledSourceOperation,
    Conflict,
    MergeOperation,
    MergeSpec,
    RunRecord,
    StackGroup,
    UploadedWorkbook,
    VerificationResult,
    WorkbookProfile,
    WorkbookSheetProfile,
)


class RecoverableExecutionIssue(Exception):
    def __init__(
        self,
        code: str,
        message: str,
        question: str,
        context: dict[str, Any],
        allowed_actions: list[str],
        phase: str = "execution",
    ) -> None:
        super().__init__(message)
        self.code = code
        self.message = message
        self.question = question
        self.context = context
        self.allowed_actions = allowed_actions
        self.phase = phase


def _source_runtime_issue(
    code: str,
    source_path: Path,
    operation: MergeOperation,
    detail: str,
) -> RecoverableExecutionIssue:
    return RecoverableExecutionIssue(
        code=code,
        message=detail,
        question=(
            f"Execution cannot safely continue with {source_path.name}. "
            "How should this source be handled?"
        ),
        context={
            "source_file": source_path.name,
            "operation_id": operation.id,
            "sheet": operation.input_sheet,
        },
        allowed_actions=[
            "retry_execution",
            "exclude_source_and_retry",
            "return_to_planning",
            "abort",
        ],
    )


def _open_source_workbook(
    source_path: Path,
    operation: MergeOperation,
    *,
    data_only: bool,
) -> Any:
    try:
        workbook = load_workbook(source_path, data_only=data_only)
    except (OSError, BadZipFile, InvalidFileException) as exc:
        raise _source_runtime_issue(
            "source_unreadable",
            source_path,
            operation,
            f"Source workbook {source_path.name!r} could not be read during execution.",
        ) from exc
    if operation.input_sheet not in workbook.sheetnames:
        workbook.close()
        raise _source_runtime_issue(
            "source_sheet_missing",
            source_path,
            operation,
            f"Source sheet {operation.input_sheet!r} disappeared before execution.",
        )
    return workbook


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def inspect_workbook(path: Path) -> WorkbookProfile:
    workbook = load_workbook(path, read_only=False, data_only=False)
    sheets = [
        WorkbookSheetProfile(
            name=worksheet.title,
            max_row=worksheet.max_row,
            max_column=worksheet.max_column,
            formula_cells=sum(
                1
                for row in worksheet.iter_rows()
                for cell in row
                if cell.data_type == "f"
            ),
        )
        for worksheet in workbook.worksheets
    ]
    workbook.close()
    return WorkbookProfile(filename=path.name, sheets=sheets)


def build_planning_evidence(template_path: Path, source_paths: list[Path]) -> dict[str, Any]:
    template_book = load_workbook(template_path, data_only=False)
    template_sheets: list[dict[str, Any]] = []
    for worksheet in template_book.worksheets:
        nonempty: list[dict[str, Any]] = []
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                nonempty.append(
                    {
                        "cell": cell.coordinate,
                        "value": str(cell.value)[:500],
                        "kind": "formula" if cell.data_type == "f" else "value",
                        "bold": bool(cell.font.bold),
                        "fill": cell.fill.fgColor.rgb if cell.fill.fill_type else None,
                    }
                )
        template_sheets.append(
            {
                "name": worksheet.title,
                "max_row": worksheet.max_row,
                "max_column": worksheet.max_column,
                "merged_ranges": [str(item) for item in worksheet.merged_cells.ranges],
                "nonempty_cells": nonempty[:400],
            }
        )
    template_book.close()

    source_summaries: list[dict[str, Any]] = []
    schema_groups: dict[str, dict[str, Any]] = {}
    for source_path in source_paths:
        workbook = load_workbook(source_path, data_only=False)
        sheets: list[dict[str, Any]] = []
        for worksheet in workbook.worksheets:
            labels = [
                {"row": row, "label": str(worksheet.cell(row, 1).value)[:300]}
                for row in range(1, worksheet.max_row + 1)
                if worksheet.cell(row, 1).value not in (None, "")
            ]
            labeled_rows = []
            for row in range(1, worksheet.max_row + 1):
                label = worksheet.cell(row, 1).value
                if label in (None, ""):
                    continue
                labeled_rows.append(
                    {
                        "row": row,
                        "label": str(label)[:300],
                        "numeric_columns": [
                            column
                            for column in range(2, worksheet.max_column + 1)
                            if _is_number(worksheet.cell(row, column).value)
                        ],
                        "nonempty_columns": [
                            column
                            for column in range(2, worksheet.max_column + 1)
                            if worksheet.cell(row, column).value not in (None, "")
                        ],
                    }
                )
            type_counts = {"number": 0, "text": 0, "formula": 0, "blank": 0}
            for row in worksheet.iter_rows():
                for cell in row:
                    if cell.value is None:
                        type_counts["blank"] += 1
                    elif cell.data_type == "f":
                        type_counts["formula"] += 1
                    elif _is_number(cell.value):
                        type_counts["number"] += 1
                    else:
                        type_counts["text"] += 1
            sheets.append(
                {
                    "name": worksheet.title,
                    "max_row": worksheet.max_row,
                    "max_column": worksheet.max_column,
                    "column_a_labels": labels[:150],
                    "labeled_rows": labeled_rows[:150],
                    "cell_type_counts": type_counts,
                }
            )
        workbook.close()
        fingerprint_payload = json.dumps(sheets, ensure_ascii=False, sort_keys=True)
        fingerprint = hashlib.sha256(fingerprint_payload.encode("utf-8")).hexdigest()
        group = schema_groups.setdefault(
            fingerprint,
            {
                "schema_fingerprint": fingerprint,
                "count": 0,
                "representative_files": [],
                "sheet_shapes": [
                    {
                        "name": sheet["name"],
                        "max_row": sheet["max_row"],
                        "max_column": sheet["max_column"],
                    }
                    for sheet in sheets
                ],
            },
        )
        group["count"] += 1
        if len(group["representative_files"]) < 3:
            group["representative_files"].append(source_path.name)
        # Bound model context while retaining representatives for every observed
        # schema variation (up to a defensive global cap).
        if group["count"] <= 2 and len(source_summaries) < 24:
            source_summaries.append(
                {
                    "filename": source_path.name,
                    "schema_fingerprint": fingerprint,
                    "sheets": sheets,
                }
            )

    return {
        "template_filename": template_path.name,
        "template_sheets": template_sheets,
        "source_workbooks": source_summaries,
        "source_population": {
            "total_workbooks": len(source_paths),
            "representatives_in_prompt": len(source_summaries),
            "schema_groups": list(schema_groups.values()),
        },
        "executor_contract": {
            "supported_modes": ["add", "concatenate"],
            "configuration_driven": True,
            "column_numbers_are_one_based": True,
            "column_alignment": {
                "default": "auto",
                "modes": ["auto", "header_path", "position"],
                "behavior": "map source columns to template columns by hierarchical merged-header paths; inserted or reordered columns do not shift values",
            },
            "add": {
                "required": ["sheet", "row_keys or data_start_row", "value_columns"],
                "alignments": ["row_key", "position"],
                "placements": ["in_place", "stack"],
                "behavior": "sum configured cells across sources; unexpected text blocks for human review",
            },
            "concatenate": {
                "required": ["sheet", "data_start_row", "end_marker_prefix", "column_count", "style_template_row", "stack_group", "row_filter"],
                "behavior": "copy qualifying rows in source-file order and freeze displayed formula values",
            },
            "row_filter": {
                "fields": ["exclude_prefixes", "exclude_exact_values", "exclude_contains", "exclude_regexes"],
                "behavior": "classify template examples, placeholders, totals, and other non-output body rows without built-in language assumptions",
            },
            "stack_groups": "Define a shared output body for ordered concatenate/add fragments and an optional retained marker row.",
            "formula_policy": "freeze_displayed_value",
            "blank_numeric_policy": "zero",
        },
    }


def _row_for_label(worksheet: Any, label: str, key_column: int = 1) -> int | None:
    normalized = _normalized_key(label)
    matches = [
        row
        for row in range(1, worksheet.max_row + 1)
        if _normalized_key(worksheet.cell(row, key_column).value) == normalized
    ]
    if len(matches) > 1:
        raise ValueError(
            f"Row key {label!r} is duplicated in {worksheet.title} at rows {matches}"
        )
    return matches[0] if matches else None


def _normalized_key(value: Any) -> str:
    text = str(value or "").strip().casefold()
    return re.sub(r"[\s\u3000]+", "", text)


def _marker_row(worksheet: Any, prefix: str, key_column: int = 1) -> int:
    for row in range(1, worksheet.max_row + 1):
        if str(worksheet.cell(row, key_column).value or "").strip().startswith(prefix):
            return row
    raise ValueError(f"Could not find marker {prefix!r} in {worksheet.title}")


def _selected_rows(worksheet: Any, operation: MergeOperation) -> list[tuple[str, int]]:
    key_column = operation.key_column or 1
    if operation.row_keys:
        result: list[tuple[str, int]] = []
        for key in operation.row_keys:
            row = _row_for_label(worksheet, key, key_column)
            if row is None:
                raise ValueError(f"Row key {key!r} is absent from {worksheet.title}")
            result.append((key, row))
        return result
    if operation.data_start_row is None:
        raise ValueError(f"{operation.id} requires row_keys or data_start_row")
    end_row = (
        _marker_row(worksheet, operation.end_marker_prefix, key_column)
        if operation.end_marker_prefix
        else worksheet.max_row + 1
    )
    return [
        (str(worksheet.cell(row, key_column).value or row), row)
        for row in range(operation.data_start_row, end_row)
    ]


def _normalized_header(value: Any) -> str:
    return re.sub(r"\s+", "", str(value or "")).strip().casefold()


def _merged_header_value(worksheet: Any, row: int, column: int) -> Any:
    cell = worksheet.cell(row, column)
    if cell.value not in (None, ""):
        return cell.value
    for merged in worksheet.merged_cells.ranges:
        if merged.min_row <= row <= merged.max_row and merged.min_col <= column <= merged.max_col:
            return worksheet.cell(merged.min_row, merged.min_col).value
    return None


def _header_rows(
    worksheet: Any,
    before_row: int,
    key_column: int,
    max_column: int,
) -> list[int]:
    """Find structural header rows without assuming their absolute positions."""
    result: list[int] = []
    for row in range(1, max(1, before_row)):
        key_value = worksheet.cell(row, key_column).value
        values = [
            _merged_header_value(worksheet, row, column)
            for column in range(1, min(max_column, worksheet.max_column) + 1)
        ]
        nonblank = [value for value in values if value not in (None, "")]
        text_count = sum(isinstance(value, str) for value in nonblank)
        if (key_value in (None, "") and any(values[1:])) or text_count >= 2:
            result.append(row)
    return result


def _column_signature(
    worksheet: Any,
    column: int,
    header_rows: list[int],
) -> tuple[str, ...]:
    signature: list[str] = []
    for row in header_rows:
        value = _normalized_header(_merged_header_value(worksheet, row, column))
        if value and (not signature or signature[-1] != value):
            signature.append(value)
    return tuple(signature)


def _first_source_body_row(worksheet: Any, operation: MergeOperation) -> int:
    if operation.row_keys:
        rows = [
            row
            for key in operation.row_keys
            if (row := _row_for_label(worksheet, key, operation.key_column or 1)) is not None
        ]
        return min(rows) if rows else worksheet.max_row + 1
    start = operation.data_start_row or 1
    end = (
        _marker_row(worksheet, operation.end_marker_prefix, operation.key_column or 1)
        if operation.end_marker_prefix
        else worksheet.max_row + 1
    )
    key_column = operation.key_column or 1
    for row in range(start, end):
        label = worksheet.cell(row, key_column).value
        values = [worksheet.cell(row, column).value for column in range(1, worksheet.max_column + 1)]
        if _is_included_row(label, values, operation):
            return row
    return end


def _source_column_mapping(
    template_sheet: Any,
    source_sheet: Any,
    operation: MergeOperation,
    target_columns: Iterable[int],
) -> dict[int, int]:
    """Map target columns to source columns using hierarchical header meaning."""
    columns = list(target_columns)
    if operation.column_alignment == "position":
        return {column: column for column in columns}

    key_column = operation.key_column or 1
    if operation.mode == "add":
        target_rows = _selected_rows(template_sheet, operation)
        target_boundary = min(row for _, row in target_rows)
    else:
        target_boundary = operation.data_start_row or 1
    source_boundary = _first_source_body_row(source_sheet, operation)
    target_headers = _header_rows(
        template_sheet, target_boundary, key_column, template_sheet.max_column
    )
    source_headers = _header_rows(
        source_sheet, source_boundary, key_column, source_sheet.max_column
    )
    source_signatures: dict[tuple[str, ...], list[int]] = {}
    for column in range(1, source_sheet.max_column + 1):
        signature = _column_signature(source_sheet, column, source_headers)
        if signature:
            source_signatures.setdefault(signature, []).append(column)

    mapping: dict[int, int] = {}
    for target_column in columns:
        if target_column == key_column:
            mapping[target_column] = key_column
            continue
        signature = _column_signature(template_sheet, target_column, target_headers)
        candidates = source_signatures.get(signature, []) if signature else []
        if len(candidates) == 1:
            mapping[target_column] = candidates[0]
        elif operation.column_alignment == "auto" and not signature and target_column <= source_sheet.max_column:
            mapping[target_column] = target_column
        elif not candidates:
            raise ValueError(
                f"No source column matches target column {target_column} header path {signature!r}"
            )
        else:
            raise ValueError(
                f"Multiple source columns match target column {target_column} header path {signature!r}: {candidates}"
            )
    return mapping


def compile_merge_plan(
    template_path: Path,
    sources: list[UploadedWorkbook],
    spec: MergeSpec,
) -> CompiledMergePlan:
    """Resolve all structural choices before approval so execution does no matching."""
    template = load_workbook(template_path, data_only=False)
    compiled_operations: list[CompiledOperation] = []
    try:
        for operation in spec.operations:
            target_sheet = template[operation.sheet]
            target_columns = list(
                operation.value_columns
                if operation.mode == "add"
                else range(1, (operation.column_count or 0) + 1)
            )
            target_rows = (
                _selected_rows(target_sheet, operation)
                if operation.mode == "add"
                else []
            )
            source_entries: list[CompiledSourceOperation] = []
            for source in sources:
                workbook = load_workbook(Path(source.stored_path), data_only=True)
                try:
                    if operation.input_sheet not in workbook.sheetnames:
                        continue
                    worksheet = workbook[operation.input_sheet]
                    mapping = _source_column_mapping(
                        target_sheet, worksheet, operation, target_columns
                    )
                    target_boundary = (
                        min(row for _, row in target_rows)
                        if target_rows
                        else operation.data_start_row or 1
                    )
                    header_rows = _header_rows(
                        target_sheet,
                        target_boundary,
                        operation.key_column or 1,
                        target_sheet.max_column,
                    )
                    columns = []
                    for target_column, source_column in mapping.items():
                        signature = list(
                            _column_signature(target_sheet, target_column, header_rows)
                        )
                        if target_column == (operation.key_column or 1):
                            matched_by = "key_column"
                        elif operation.column_alignment == "position" or not signature:
                            matched_by = "position"
                        else:
                            matched_by = "header_path"
                        columns.append(
                            CompiledColumnMapping(
                                target_column=target_column,
                                source_column=source_column,
                                header_path=signature,
                                matched_by=matched_by,
                            )
                        )
                    rows: list[CompiledRowMapping] = []
                    if operation.mode == "add":
                        for key, target_row in target_rows:
                            source_row = (
                                _row_for_label(
                                    worksheet, key, operation.key_column or 1
                                )
                                if operation.alignment == "row_key"
                                else target_row
                            )
                            if source_row is None:
                                raise ValueError(
                                    f"Row key {key!r} is absent from {worksheet.title}"
                                )
                            rows.append(
                                CompiledRowMapping(
                                    source_row=source_row,
                                    target_row=target_row,
                                    row_key=key,
                                )
                            )
                    else:
                        end_row = _marker_row(
                            worksheet,
                            operation.end_marker_prefix or "",
                            operation.key_column or 1,
                        )
                        for row in range(operation.data_start_row or 1, end_row):
                            values = [
                                worksheet.cell(row, mapping[column]).value
                                for column in target_columns
                            ]
                            label = values[(operation.key_column or 1) - 1]
                            if _is_included_row(label, values, operation):
                                rows.append(
                                    CompiledRowMapping(
                                        source_row=row,
                                        row_key=str(label).strip(),
                                    )
                                )
                    source_entries.append(
                        CompiledSourceOperation(
                            source_id=source.id,
                            source_file=source.filename,
                            source_sheet=operation.input_sheet,
                            target_sheet=operation.sheet,
                            columns=columns,
                            rows=rows,
                        )
                    )
                finally:
                    workbook.close()
            compiled_operations.append(
                CompiledOperation(
                    operation_id=operation.id,
                    mode=operation.mode,
                    sources=source_entries,
                )
            )
    finally:
        template.close()
    return CompiledMergePlan(spec_hash=spec.digest(), operations=compiled_operations)


def validate_merge_spec(spec: MergeSpec, template_path: Path) -> MergeSpec:
    if not spec.operations:
        raise ValueError("Plan must contain at least one operation")
    ids = [item.id for item in spec.operations]
    if len(ids) != len(set(ids)):
        raise ValueError("Operation ids must be unique")
    group_map = {group.id: group for group in spec.stack_groups}
    if len(group_map) != len(spec.stack_groups):
        raise ValueError("Stack group ids must be unique")

    workbook = load_workbook(template_path, data_only=False)
    try:
        for group in spec.stack_groups:
            if group.sheet not in workbook.sheetnames:
                raise ValueError(f"Stack group {group.id!r} references missing sheet {group.sheet!r}")
            worksheet = workbook[group.sheet]
            if group.column_count > worksheet.max_column:
                raise ValueError(f"Stack group {group.id!r} exceeds template width")
            if group.end_marker_prefix:
                _marker_row(worksheet, group.end_marker_prefix)

        in_place_cells: set[tuple[str, str, int]] = set()
        for operation in spec.operations:
            if operation.sheet not in workbook.sheetnames:
                raise ValueError(f"Operation {operation.id!r} references missing target sheet {operation.sheet!r}")
            worksheet = workbook[operation.sheet]
            if operation.mode == "add":
                if not operation.value_columns:
                    raise ValueError(f"Add operation {operation.id!r} requires value_columns")
                if operation.key_column is None and operation.alignment == "row_key":
                    raise ValueError(f"Add operation {operation.id!r} requires key_column for row_key alignment")
                selected = _selected_rows(worksheet, operation)
                if max(operation.value_columns) > worksheet.max_column:
                    raise ValueError(f"Operation {operation.id!r} references a column outside the template")
                if operation.placement == "in_place":
                    for key, _ in selected:
                        for column in operation.value_columns:
                            signature = (operation.sheet, key, column)
                            if signature in in_place_cells:
                                raise ValueError("Two in-place add operations write the same target cell")
                            in_place_cells.add(signature)
            else:
                required = [operation.data_start_row, operation.end_marker_prefix, operation.column_count, operation.style_template_row]
                if any(item is None for item in required):
                    raise ValueError(f"Concatenate operation {operation.id!r} is missing range/style parameters")
                if operation.placement != "stack":
                    raise ValueError("Concatenate operations must use a stack group")
                if operation.column_count and operation.column_count > worksheet.max_column:
                    raise ValueError(f"Operation {operation.id!r} exceeds template width")
                if operation.style_template_row and operation.style_template_row > worksheet.max_row:
                    raise ValueError(f"Operation {operation.id!r} style row is outside the template")

            if operation.placement == "stack":
                if not operation.stack_group or operation.stack_group not in group_map:
                    raise ValueError(f"Operation {operation.id!r} references an unknown stack group")
                if group_map[operation.stack_group].sheet != operation.sheet:
                    raise ValueError(f"Operation {operation.id!r} and its stack group target different sheets")

        for operation in spec.operations:
            for pattern in operation.row_filter.exclude_regexes:
                try:
                    re.compile(pattern)
                except re.error as exc:
                    raise ValueError(
                        f"Operation {operation.id!r} has invalid row-filter regex {pattern!r}"
                    ) from exc

        for concatenate in (item for item in spec.operations if item.mode == "concatenate"):
            stacked_keys = {
                key
                for item in spec.operations
                if item.mode == "add" and item.stack_group == concatenate.stack_group
                for key in item.row_keys
            }
            uncovered_stacked_keys = [
                key for key in stacked_keys if not concatenate.row_filter.matches(key)
            ]
            if uncovered_stacked_keys:
                raise ValueError(
                    f"Concatenate operation {concatenate.id!r} must exclude stacked add row keys to prevent duplicates"
                )
            worksheet = workbook[concatenate.sheet]
            concatenate_end = _marker_row(
                worksheet,
                concatenate.end_marker_prefix or "",
                concatenate.key_column or 1,
            )
            group = group_map[concatenate.stack_group or ""]
            template_end = (
                _marker_row(worksheet, group.end_marker_prefix, concatenate.key_column or 1)
                if group.end_marker_prefix
                else concatenate_end
            )
            if concatenate_end != template_end:
                raise ValueError(
                    f"Concatenate operation {concatenate.id!r} ends before the full stack-group body"
                )
            unclassified: list[str] = []
            for row in range(concatenate.data_start_row or 1, template_end):
                label = str(
                    worksheet.cell(row, concatenate.key_column or 1).value or ""
                ).strip()
                if not label or label in stacked_keys or concatenate.row_filter.matches(label):
                    continue
                unclassified.append(label)
            if unclassified:
                raise ValueError(
                    f"Concatenate operation {concatenate.id!r} leaves template body rows unclassified: {unclassified}"
                )

        concatenate_signatures: set[tuple[Any, ...]] = set()
        for operation in (item for item in spec.operations if item.mode == "concatenate"):
            signature = (
                operation.input_sheet,
                operation.sheet,
                operation.data_start_row,
                operation.end_marker_prefix,
                operation.column_count,
                operation.stack_group,
            )
            if signature in concatenate_signatures:
                raise ValueError(
                    "Duplicate concatenate operations would copy the same rows from every source more than once"
                )
            concatenate_signatures.add(signature)

        if not spec.guideline_citations:
            raise ValueError("The model plan must cite at least one template guideline cell")
        for citation in spec.guideline_citations:
            match = re.fullmatch(r"(.+)!([A-Z]+[1-9][0-9]*)", citation)
            if not match or match.group(1) not in workbook.sheetnames:
                raise ValueError(f"Invalid guideline citation {citation!r}")
            if workbook[match.group(1)][match.group(2)].value is None:
                raise ValueError(f"Guideline citation {citation!r} points to a blank cell")
    finally:
        workbook.close()
    return spec


def _cell_error(value: Any) -> bool:
    return isinstance(value, str) and value.startswith("#")


def _is_number(value: Any) -> bool:
    return isinstance(value, (int, float)) and not isinstance(value, bool)


def _conflict_id(operation_id: str, filename: str, sheet: str, cell: str) -> str:
    value = f"{operation_id}|{filename}|{sheet}|{cell}".encode()
    return "conflict-" + hashlib.sha256(value).hexdigest()[:12]


def detect_conflicts(
    template_path: Path, sources: list[UploadedWorkbook], spec: MergeSpec
) -> list[Conflict]:
    conflicts: list[Conflict] = []
    template = load_workbook(template_path, data_only=False)
    template_sheets = set(template.sheetnames)

    for operation in spec.operations:
        if operation.sheet not in template_sheets:
            conflicts.append(
                Conflict(
                    id=_conflict_id(operation.id, template_path.name, operation.sheet, "template"),
                    type="missing_sheet",
                    source_file=template_path.name,
                    sheet=operation.sheet,
                    expected="target template sheet",
                    actual=None,
                    message=f"Template is missing target sheet {operation.sheet!r}.",
                    recommended_action="abort",
                    allowed_actions=["abort"],
                )
            )

    seen_hashes: dict[str, UploadedWorkbook] = {}
    for source in sources:
        source_path = Path(source.stored_path)
        duplicate = seen_hashes.get(source.sha256)
        if duplicate is not None:
            conflicts.append(
                Conflict(
                    id=_conflict_id("duplicate-source", source.id, source.filename, "file"),
                    type="duplicate_source",
                    source_id=source.id,
                    source_file=source.filename,
                    sheet="",
                    expected="a unique source workbook",
                    actual=duplicate.filename,
                    message=(
                        f"{source.filename!r} has the same content as {duplicate.filename!r}; "
                        "including both would double count it."
                    ),
                    recommended_action="exclude_source",
                    allowed_actions=["exclude_source", "abort"],
                )
            )
            continue
        seen_hashes[source.sha256] = source
        workbook = load_workbook(source_path, data_only=True)
        for operation in spec.operations:
            sheet = operation.input_sheet
            if sheet not in workbook.sheetnames:
                conflicts.append(
                    Conflict(
                        id=_conflict_id(operation.id, source_path.name, sheet, "missing"),
                        type="missing_sheet",
                        source_id=source.id,
                        source_file=source_path.name,
                        sheet=sheet,
                        expected="configured source sheet",
                        actual=None,
                        message=f"Source sheet {sheet!r} required by {operation.id!r} is missing.",
                        recommended_action="exclude_source",
                        allowed_actions=["exclude_source", "abort"],
                    )
                )
                continue
            worksheet = workbook[sheet]
            rows: list[tuple[str, int]] | None = None
            if operation.mode == "add":
                try:
                    rows = _selected_rows(worksheet, operation)
                except ValueError as exc:
                    duplicate_key = "duplicated" in str(exc)
                    conflicts.append(
                        Conflict(
                            id=_conflict_id(operation.id, source.id, sheet, "rows"),
                            type="duplicate_row_key" if duplicate_key else "missing_row_key",
                            source_id=source.id,
                            source_file=source.filename,
                            sheet=sheet,
                            expected="one unique configured input row",
                            actual=None,
                            message=str(exc),
                            recommended_action="exclude_source",
                            allowed_actions=["exclude_source", "abort"],
                        )
                    )
                    continue
            target_columns = (
                operation.value_columns
                if operation.mode == "add"
                else range(1, (operation.column_count or 0) + 1)
            )
            try:
                column_mapping = _source_column_mapping(
                    template[operation.sheet], worksheet, operation, target_columns
                )
            except ValueError as exc:
                conflicts.append(
                    Conflict(
                        id=_conflict_id(operation.id, source_path.name, sheet, "columns"),
                        type="schema_mismatch",
                        source_id=source.id,
                        source_file=source_path.name,
                        sheet=sheet,
                        expected="unique semantic column headers matching the template",
                        actual=None,
                        message=f"{operation.id}: {exc}",
                        recommended_action="exclude_source",
                        allowed_actions=["exclude_source", "abort"],
                    )
                )
                continue
            if operation.mode == "concatenate":
                if operation.end_marker_prefix:
                    try:
                        _marker_row(worksheet, operation.end_marker_prefix, operation.key_column or 1)
                    except ValueError:
                        conflicts.append(
                            Conflict(
                                id=_conflict_id(operation.id, source_path.name, sheet, "marker"),
                                type="schema_mismatch",
                                source_id=source.id,
                                source_file=source_path.name,
                                sheet=sheet,
                                expected=f"end marker {operation.end_marker_prefix!r}",
                                actual=None,
                                message=f"Could not find the configured concatenate boundary for {operation.id!r}.",
                                recommended_action="exclude_source",
                                allowed_actions=["exclude_source", "abort"],
                            )
                        )
                continue

            assert rows is not None
            for key, row in rows:
                for column in operation.value_columns:
                    cell = worksheet.cell(row, column_mapping[column])
                    if cell.value is None or _is_number(cell.value):
                        continue
                    kind = "formula_error" if _cell_error(cell.value) else "numeric_text_mismatch"
                    conflicts.append(
                        Conflict(
                            id=_conflict_id(operation.id, source_path.name, sheet, cell.coordinate),
                            type=kind,
                            source_id=source.id,
                            source_file=source_path.name,
                            sheet=sheet,
                            cell=cell.coordinate,
                            row_key=key,
                            expected="number or blank",
                            actual=cell.value,
                            message=f"{operation.id}: {cell.coordinate} contains {cell.value!r}, which cannot be added safely.",
                            recommended_action="treat_as_zero" if cell.value == "/" else "skip_cell",
                            allowed_actions=["treat_as_zero", "keep_marker", "skip_cell", "exclude_source", "abort"],
                        )
                    )
        workbook.close()
    template.close()
    return conflicts


def _source_excluded(run: RunRecord, source: UploadedWorkbook) -> bool:
    return source.id in run.excluded_sources or source.filename in run.excluded_sources or any(
        (conflict.source_id == source.id or conflict.source_file == source.filename)
        and conflict.resolution == "exclude_source"
        for conflict in run.conflicts
    )


def _resolve_add_value(run: RunRecord, source: UploadedWorkbook, sheet: str, coordinate: str, value: Any) -> tuple[float | None, str | None]:
    if value is None:
        return None, None
    if _is_number(value):
        return float(value), None
    conflict = next(
        (
            item
            for item in run.conflicts
            if (item.source_id == source.id or item.source_file == source.filename)
            and item.sheet == sheet
            and item.cell == coordinate
        ),
        None,
    )
    if conflict is None or not conflict.resolved:
        raise RecoverableExecutionIssue(
            code="unresolved_runtime_value",
            message=f"A new unresolved value appeared at {source.filename}:{sheet}!{coordinate}.",
            question=f"The source value at {source.filename}:{sheet}!{coordinate} changed after planning. How should execution continue?",
            context={"source_id": source.id, "source_file": source.filename, "sheet": sheet, "cell": coordinate},
            allowed_actions=[
                "retry_execution",
                "exclude_source_and_retry",
                "return_to_planning",
                "abort",
            ],
        )
    if conflict.resolution == "treat_as_zero":
        return 0.0, None
    if conflict.resolution == "keep_marker":
        return None, str(value)
    if conflict.resolution in {"skip_cell", "exclude_source"}:
        return None, None
    raise ValueError("Merge aborted by conflict resolution")


@dataclass
class RowStyle:
    styles: list[Any]
    height: float | None


def _capture_row_style(worksheet: Any, row: int, max_column: int) -> RowStyle:
    return RowStyle(
        styles=[copy(worksheet.cell(row, column)._style) for column in range(1, max_column + 1)],
        height=worksheet.row_dimensions[row].height,
    )


def _apply_row_style(worksheet: Any, row: int, snapshot: RowStyle) -> None:
    for column, style in enumerate(snapshot.styles, start=1):
        worksheet.cell(row, column)._style = copy(style)
    worksheet.row_dimensions[row].height = snapshot.height


def _clean_number(value: float) -> int | float:
    return int(value) if value.is_integer() else value


def _is_included_row(label: Any, row_values: Iterable[Any], operation: MergeOperation) -> bool:
    text = str(label or "").strip()
    if not text or operation.row_filter.matches(text):
        return False
    return any(value not in (None, "") for value in row_values)


def _compiled_source(
    run: RunRecord, operation_id: str, source_id: str
) -> CompiledSourceOperation | None:
    if run.compiled_plan is None:
        return None
    operation = next(
        (
            item
            for item in run.compiled_plan.operations
            if item.operation_id == operation_id
        ),
        None,
    )
    if operation is None:
        return None
    return next(
        (item for item in operation.sources if item.source_id == source_id), None
    )


BatchProgressCallback = Callable[[str, int, int, int], None]


def _source_batches(
    sources: list[UploadedWorkbook], batch_size: int
) -> Iterator[tuple[int, int, list[UploadedWorkbook]]]:
    total = (len(sources) + batch_size - 1) // batch_size
    for start in range(0, len(sources), batch_size):
        yield start // batch_size + 1, total, sources[start : start + batch_size]


def _sum_operation_rows(
    run: RunRecord,
    operation: MergeOperation,
    output_sheet: Any,
    sources: list[UploadedWorkbook],
    width: int,
    batch_size: int,
    progress_callback: BatchProgressCallback | None = None,
) -> list[tuple[list[Any], RowStyle]]:
    target_rows = _selected_rows(output_sheet, operation)
    active_sources = [source for source in sources if not _source_excluded(run, source)]
    row_state: dict[tuple[str, int], dict[str, Any]] = {}
    for key, target_row in target_rows:
        row_state[(key, target_row)] = {
            "values": [output_sheet.cell(target_row, column).value for column in range(1, width + 1)],
            "style": _capture_row_style(output_sheet, target_row, width),
            "columns": {
                column: {"total": 0.0, "numeric_seen": False, "marker": None}
                for column in operation.value_columns
            },
        }

    for batch_index, batch_count, batch in _source_batches(active_sources, batch_size):
        books = {
            source.id: _open_source_workbook(
                Path(source.stored_path), operation, data_only=True
            )
            for source in batch
        }
        try:
            for source in batch:
                source_sheet = books[source.id][operation.input_sheet]
                compiled = _compiled_source(run, operation.id, source.id)
                if compiled is None:
                    source_path = Path(source.stored_path)
                    raise _source_runtime_issue(
                        "compiled_mapping_missing", source_path, operation,
                        f"No reviewed mapping exists for {source.filename!r} and {operation.id!r}.",
                    )
                for key, target_row in target_rows:
                    state = row_state[(key, target_row)]
                    row_entry = next(
                        (item for item in compiled.rows if _normalized_key(item.row_key) == _normalized_key(key)),
                        None,
                    )
                    for column in operation.value_columns:
                        source_path = Path(source.stored_path)
                        column_entry = next(
                            (item for item in compiled.columns if item.target_column == column), None
                        )
                        if column_entry is None or row_entry is None:
                            raise _source_runtime_issue(
                                "compiled_mapping_incomplete", source_path, operation,
                                f"The reviewed mapping is incomplete for row {key!r}, column {column}.",
                            )
                        cell = source_sheet.cell(row_entry.source_row, column_entry.source_column)
                        number, source_marker = _resolve_add_value(
                            run, source, operation.input_sheet, cell.coordinate, cell.value
                        )
                        accumulator = state["columns"][column]
                        if number is not None:
                            accumulator["total"] += number
                            accumulator["numeric_seen"] = True
                        if source_marker is not None:
                            accumulator["marker"] = source_marker
        finally:
            for workbook in books.values():
                workbook.close()
        if progress_callback:
            progress_callback(
                operation.id,
                batch_index,
                batch_count,
                min(batch_index * batch_size, len(active_sources)),
            )

    output: list[tuple[list[Any], RowStyle]] = []
    for key, target_row in target_rows:
        state = row_state[(key, target_row)]
        for column, accumulator in state["columns"].items():
            state["values"][column - 1] = (
                accumulator["marker"]
                if accumulator["marker"] is not None
                else _clean_number(accumulator["total"])
                if accumulator["numeric_seen"]
                else None
            )
        output.append((state["values"], state["style"]))
    return output


def _concatenate_rows(
    run: RunRecord,
    operation: MergeOperation,
    output_sheet: Any,
    sources: list[UploadedWorkbook],
    audit: dict[str, Any],
    width: int,
    batch_size: int,
    progress_callback: BatchProgressCallback | None = None,
) -> Iterable[tuple[list[Any], RowStyle]]:
    style = _capture_row_style(output_sheet, operation.style_template_row or 1, width)
    active_sources = [source for source in sources if not _source_excluded(run, source)]

    def rows() -> Iterator[tuple[list[Any], RowStyle]]:
        for batch_index, batch_count, batch in _source_batches(active_sources, batch_size):
            for source in batch:
                source_path = Path(source.stored_path)
                compiled = _compiled_source(run, operation.id, source.id)
                if compiled is None:
                    raise _source_runtime_issue(
                        "compiled_mapping_missing", source_path, operation,
                        f"No reviewed mapping exists for {source.filename!r} and {operation.id!r}.",
                    )
                values_book = _open_source_workbook(source_path, operation, data_only=True)
                formulas_book = _open_source_workbook(source_path, operation, data_only=False)
                try:
                    values_sheet = values_book[operation.input_sheet]
                    formulas_sheet = formulas_book[operation.input_sheet]
                    column_mapping = {
                        item.target_column: item.source_column for item in compiled.columns
                    }
                    for row_entry in compiled.rows:
                        row = row_entry.source_row
                        values = [
                            values_sheet.cell(row, column_mapping[column]).value
                            for column in range(1, width + 1)
                        ]
                        yield values, style
                        for column in range(1, width + 1):
                            formula_cell = formulas_sheet.cell(row, column_mapping[column])
                            if formula_cell.data_type == "f":
                                audit["frozen_formulas"].append(
                                    {
                                        "source_id": source.id,
                                        "source_file": source.filename,
                                        "sheet": operation.input_sheet,
                                        "cell": formula_cell.coordinate,
                                        "formula": formula_cell.value,
                                        "displayed_value": values[column - 1],
                                    }
                                )
                finally:
                    values_book.close()
                    formulas_book.close()
            if progress_callback:
                progress_callback(
                    operation.id,
                    batch_index,
                    batch_count,
                    min(batch_index * batch_size, len(active_sources)),
                )
    return rows()


def _canonical_value(value: Any) -> dict[str, Any]:
    if value is None:
        return {"type": "blank", "value": None}
    if isinstance(value, bool):
        return {"type": "boolean", "value": value}
    if _is_number(value):
        return {"type": "number", "value": format(float(value), ".15g")}
    if isinstance(value, (date, datetime)):
        return {"type": "date", "value": value.isoformat()}
    return {"type": "text", "value": str(value)}


def _row_fingerprint(values: list[Any]) -> str:
    payload = json.dumps(
        [_canonical_value(value) for value in values],
        ensure_ascii=False,
        sort_keys=True,
    )
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()


def build_reconciliation_manifest(run: RunRecord, spec: MergeSpec) -> dict[str, Any]:
    """Independently recompute expected writes from the compiled plan and sources."""
    if run.template is None or run.compiled_plan is None:
        raise ValueError("A compiled plan is required for reconciliation")
    template = load_workbook(Path(run.template.stored_path), data_only=False)
    manifest: dict[str, Any] = {"cells": [], "rows": [], "source_contributions": 0}
    source_cell_cache: dict[
        tuple[str, str], dict[tuple[str, int], tuple[str, Any]]
    ] = {}

    def source_cells(
        operation: MergeOperation,
        source: UploadedWorkbook,
        compiled: CompiledSourceOperation,
    ) -> dict[tuple[str, int], tuple[str, Any]]:
        cache_key = (operation.id, source.id)
        if cache_key in source_cell_cache:
            return source_cell_cache[cache_key]
        cells: dict[tuple[str, int], tuple[str, Any]] = {}
        value_book = load_workbook(Path(source.stored_path), data_only=True)
        try:
            sheet = value_book[operation.input_sheet]
            for row_entry in compiled.rows:
                for column_entry in compiled.columns:
                    if column_entry.target_column not in operation.value_columns:
                        continue
                    cell = sheet.cell(
                        row_entry.source_row, column_entry.source_column
                    )
                    cells[
                        (_normalized_key(row_entry.row_key), column_entry.target_column)
                    ] = (cell.coordinate, cell.value)
        finally:
            value_book.close()
        source_cell_cache[cache_key] = cells
        return cells

    def expected_add_row(operation: MergeOperation, key: str, target_row: int, width: int) -> tuple[list[Any], list[dict[str, Any]]]:
        target_sheet = template[operation.sheet]
        values = [target_sheet.cell(target_row, column).value for column in range(1, width + 1)]
        contributions: list[dict[str, Any]] = []
        for target_column in operation.value_columns:
            total = 0.0
            numeric_seen = False
            marker: str | None = None
            for source in run.sources:
                if _source_excluded(run, source):
                    continue
                compiled = _compiled_source(run, operation.id, source.id)
                if compiled is None:
                    continue
                column_entry = next(item for item in compiled.columns if item.target_column == target_column)
                row_entry = next(
                    item
                    for item in compiled.rows
                    if _normalized_key(item.row_key) == _normalized_key(key)
                )
                cell_coordinate, cell_value = source_cells(
                    operation, source, compiled
                )[(_normalized_key(row_entry.row_key), target_column)]
                number, source_marker = _resolve_add_value(
                    run, source, operation.input_sheet, cell_coordinate, cell_value
                )
                contributions.append(
                    {
                        "operation_id": operation.id,
                        "source_id": source.id,
                        "source_file": source.filename,
                        "source_cell": f"{operation.input_sheet}!{cell_coordinate}",
                        "target_sheet": operation.sheet,
                        "target_column": target_column,
                        "row_key": key,
                        "value": _canonical_value(cell_value),
                    }
                )
                if number is not None:
                    total += number
                    numeric_seen = True
                if source_marker is not None:
                    marker = source_marker
            values[target_column - 1] = marker if marker is not None else (
                _clean_number(total) if numeric_seen else None
            )
        return values, contributions

    try:
        for operation in (
            item
            for item in spec.operations
            if item.mode == "add" and item.placement == "in_place"
        ):
            width = template[operation.sheet].max_column
            for key, target_row in _selected_rows(template[operation.sheet], operation):
                values, contributions = expected_add_row(operation, key, target_row, width)
                for column in operation.value_columns:
                    manifest["cells"].append(
                        {
                            "operation_id": operation.id,
                            "sheet": operation.sheet,
                            "row": target_row,
                            "column": column,
                            "expected": _canonical_value(values[column - 1]),
                            "contributions": [
                                item
                                for item in contributions
                                if item["target_column"] == column
                            ],
                        }
                    )
                manifest["source_contributions"] += len(contributions)

        for group in spec.stack_groups:
            current_row = group.start_row
            for operation in sorted(
                (item for item in spec.operations if item.stack_group == group.id),
                key=lambda item: (item.stack_order, item.id),
            ):
                if operation.mode == "concatenate":
                    for source in run.sources:
                        if _source_excluded(run, source):
                            continue
                        compiled = _compiled_source(run, operation.id, source.id)
                        if compiled is None:
                            continue
                        column_map = {
                            item.target_column: item.source_column
                            for item in compiled.columns
                        }
                        value_book = load_workbook(Path(source.stored_path), data_only=True)
                        try:
                            sheet = value_book[operation.input_sheet]
                            for row_entry in compiled.rows:
                                values = [
                                    sheet.cell(row_entry.source_row, column_map[column]).value
                                    for column in range(1, group.column_count + 1)
                                ]
                                manifest["rows"].append(
                                    {
                                        "operation_id": operation.id,
                                        "sheet": group.sheet,
                                        "row": current_row,
                                        "width": group.column_count,
                                        "expected_fingerprint": _row_fingerprint(values),
                                        "source_id": source.id,
                                        "source_file": source.filename,
                                        "source_row": row_entry.source_row,
                                        "row_key": row_entry.row_key,
                                    }
                                )
                                manifest["source_contributions"] += group.column_count
                                current_row += 1
                        finally:
                            value_book.close()
                else:
                    target_rows = _selected_rows(template[operation.sheet], operation)
                    for key, target_row in target_rows:
                        values, contributions = expected_add_row(
                            operation, key, target_row, group.column_count
                        )
                        manifest["rows"].append(
                            {
                                "operation_id": operation.id,
                                "sheet": group.sheet,
                                "row": current_row,
                                "width": group.column_count,
                                "expected_fingerprint": _row_fingerprint(values),
                                "row_key": key,
                                "contributions": contributions,
                            }
                        )
                        manifest["source_contributions"] += len(contributions)
                        current_row += 1
    finally:
        template.close()
    manifest["expected_cell_count"] = len(manifest["cells"])
    manifest["expected_row_count"] = len(manifest["rows"])
    return manifest


def execute_merge(
    run: RunRecord,
    output_path: Path,
    audit_path: Path,
    supplied_spec: MergeSpec | None = None,
    *,
    progress_callback: BatchProgressCallback | None = None,
) -> VerificationResult:
    if run.template is None or run.spec is None:
        raise ValueError("Run is missing template or specification")
    spec = supplied_spec or run.spec
    if spec.digest() != run.spec_hash or run.approved_spec_hash != run.spec_hash:
        raise ValueError("Executor configuration does not match the approved plan hash")
    unresolved = [item for item in run.conflicts if item.severity == "blocking" and not item.resolved]
    if unresolved:
        raise ValueError(f"Run has {len(unresolved)} unresolved blocking conflicts")

    sources = list(run.sources)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(Path(run.template.stored_path), output_path)
    output_book = load_workbook(output_path, data_only=False)
    audit: dict[str, Any] = {
        "run_id": run.id,
        "template": run.template.model_dump(),
        "sources": [source.model_dump() for source in run.sources],
        "spec": spec.model_dump(),
        "spec_hash": spec.digest(),
        "compiled_plan": run.compiled_plan.model_dump() if run.compiled_plan else None,
        "compiled_plan_hash": run.compiled_plan.digest() if run.compiled_plan else None,
        "resolutions": [item.model_dump() for item in run.conflicts if item.resolution],
        "frozen_formulas": [],
        "operations": [],
        "batch_execution": {
            "batch_size": run.batch_size,
            "active_source_count": sum(
                1 for source in sources if not _source_excluded(run, source)
            ),
        },
    }

    for operation in (item for item in spec.operations if item.mode == "add" and item.placement == "in_place"):
        worksheet = output_book[operation.sheet]
        rows = _sum_operation_rows(
            run,
            operation,
            worksheet,
            sources,
            worksheet.max_column,
            run.batch_size,
            progress_callback,
        )
        for (_, target_row), (values, _) in zip(_selected_rows(worksheet, operation), rows, strict=True):
            for column in operation.value_columns:
                worksheet.cell(target_row, column).value = values[column - 1]
        audit["operations"].append({"id": operation.id, "mode": "add", "placement": "in_place", "rows": len(rows)})

    group_map = {group.id: group for group in spec.stack_groups}
    for group_id, group in group_map.items():
        worksheet = output_book[group.sheet]
        marker_style: RowStyle | None = None
        marker_value: Any = None
        marker_alignment: Any = None
        if group.end_marker_prefix:
            original_marker_row = _marker_row(worksheet, group.end_marker_prefix)
            marker_style = _capture_row_style(worksheet, original_marker_row, group.column_count)
            marker_value = worksheet.cell(original_marker_row, 1).value
            marker_alignment = copy(worksheet.cell(original_marker_row, 1).alignment)

        fragments: list[tuple[MergeOperation, Iterable[tuple[list[Any], RowStyle]]]] = []
        group_operations = sorted(
            (item for item in spec.operations if item.stack_group == group_id),
            key=lambda item: (item.stack_order, item.id),
        )
        for operation in group_operations:
            rows = (
                _sum_operation_rows(
                    run,
                    operation,
                    worksheet,
                    sources,
                    group.column_count,
                    run.batch_size,
                    progress_callback,
                )
                if operation.mode == "add"
                else _concatenate_rows(
                    run,
                    operation,
                    worksheet,
                    sources,
                    audit,
                    group.column_count,
                    run.batch_size,
                    progress_callback,
                )
            )
            fragments.append((operation, rows))

        for merged_range in list(worksheet.merged_cells.ranges):
            if merged_range.min_row >= group.start_row:
                worksheet.unmerge_cells(str(merged_range))
        for row in range(group.start_row, worksheet.max_row + 1):
            for column in range(1, group.column_count + 1):
                worksheet.cell(row, column).value = None

        current_row = group.start_row
        for operation, rows in fragments:
            row_count = 0
            for values, style in rows:
                _apply_row_style(worksheet, current_row, style)
                for column, value in enumerate(values[: group.column_count], start=1):
                    worksheet.cell(current_row, column).value = value
                current_row += 1
                row_count += 1
            audit["operations"].append(
                {"id": operation.id, "mode": operation.mode, "placement": "stack", "stack_group": group_id, "rows": row_count}
            )
        if group.retain_end_marker and marker_style is not None:
            _apply_row_style(worksheet, current_row, marker_style)
            worksheet.cell(current_row, 1).value = marker_value
            worksheet.merge_cells(
                start_row=current_row,
                start_column=1,
                end_row=current_row,
                end_column=group.column_count,
            )
            worksheet.cell(current_row, 1).alignment = marker_alignment

    output_book.save(output_path)
    output_book.close()
    manifest = build_reconciliation_manifest(run, spec)
    audit["reconciliation"] = manifest
    verification = verify_output(
        output_path,
        spec,
        manifest=manifest,
        template_path=Path(run.template.stored_path),
    )
    audit["verification"] = verification.model_dump()
    audit_path.parent.mkdir(parents=True, exist_ok=True)
    audit_path.write_text(json.dumps(audit, ensure_ascii=False, indent=2), encoding="utf-8")
    return verification


def verify_output(
    output_path: Path,
    spec: MergeSpec | None = None,
    *,
    manifest: dict[str, Any] | None = None,
    template_path: Path | None = None,
) -> VerificationResult:
    workbook = load_workbook(output_path, data_only=False)
    checks: list[dict[str, Any]] = []
    required = sorted({item.sheet for item in spec.operations}) if spec else []
    missing = [sheet for sheet in required if sheet not in workbook.sheetnames]
    checks.append({"name": "configured_sheets", "passed": not missing, "missing": missing})

    error_cells: list[str] = []
    for worksheet in workbook.worksheets:
        for row in worksheet.iter_rows():
            for cell in row:
                value = cell.value
                if _cell_error(value):
                    error_cells.append(f"{worksheet.title}!{cell.coordinate}")
    checks.append({"name": "formula_errors", "passed": not error_cells, "cells": error_cells})

    cell_mismatches: list[dict[str, Any]] = []
    row_mismatches: list[dict[str, Any]] = []
    if manifest:
        for expected in manifest.get("cells", []):
            actual = workbook[expected["sheet"]].cell(
                expected["row"], expected["column"]
            ).value
            canonical = _canonical_value(actual)
            if canonical != expected["expected"]:
                cell_mismatches.append(
                    {
                        "sheet": expected["sheet"],
                        "row": expected["row"],
                        "column": expected["column"],
                        "expected": expected["expected"],
                        "actual": canonical,
                    }
                )
        for expected in manifest.get("rows", []):
            values = [
                workbook[expected["sheet"]].cell(expected["row"], column).value
                for column in range(1, expected["width"] + 1)
            ]
            actual_fingerprint = _row_fingerprint(values)
            if actual_fingerprint != expected["expected_fingerprint"]:
                row_mismatches.append(
                    {
                        "sheet": expected["sheet"],
                        "row": expected["row"],
                        "operation_id": expected["operation_id"],
                        "expected_fingerprint": expected["expected_fingerprint"],
                        "actual_fingerprint": actual_fingerprint,
                    }
                )
    checks.append(
        {
            "name": "cell_reconciliation",
            "passed": not cell_mismatches,
            "expected_cells": len(manifest.get("cells", [])) if manifest else 0,
            "mismatches": cell_mismatches,
        }
    )
    checks.append(
        {
            "name": "row_reconciliation",
            "passed": not row_mismatches,
            "expected_rows": len(manifest.get("rows", [])) if manifest else 0,
            "mismatches": row_mismatches,
        }
    )

    untouched_mismatches: list[str] = []
    if spec and template_path:
        template = load_workbook(template_path, data_only=False)
        touched_cells = {
            (item["sheet"], item["row"], item["column"])
            for item in (manifest or {}).get("cells", [])
        }
        stack_regions = {
            group.sheet: (group.start_row, group.column_count)
            for group in spec.stack_groups
        }
        for sheet_name in template.sheetnames:
            if sheet_name not in workbook.sheetnames:
                continue
            before, after = template[sheet_name], workbook[sheet_name]
            max_row = max(before.max_row, after.max_row)
            max_column = max(before.max_column, after.max_column)
            for row in range(1, max_row + 1):
                for column in range(1, max_column + 1):
                    region = stack_regions.get(sheet_name)
                    if (sheet_name, row, column) in touched_cells or (
                        region and row >= region[0] and column <= region[1]
                    ):
                        continue
                    left, right = before.cell(row, column), after.cell(row, column)
                    if (
                        _canonical_value(left.value) != _canonical_value(right.value)
                        or left.number_format != right.number_format
                    ):
                        untouched_mismatches.append(
                            f"{sheet_name}!{right.coordinate}"
                        )
                        if len(untouched_mismatches) >= 100:
                            break
                if len(untouched_mismatches) >= 100:
                    break
        template.close()
    checks.append(
        {
            "name": "untouched_template_regions",
            "passed": not untouched_mismatches,
            "cells": untouched_mismatches,
        }
    )

    leaked_filtered_rows: list[str] = []
    if spec:
        for operation in (item for item in spec.operations if item.mode == "concatenate"):
            worksheet = workbook[operation.sheet]
            group = next(
                item for item in spec.stack_groups if item.id == operation.stack_group
            )
            body_end = (
                _marker_row(worksheet, group.end_marker_prefix, operation.key_column or 1)
                if group.end_marker_prefix
                else worksheet.max_row + 1
            )
            stacked_keys = {
                key
                for item in spec.operations
                if item.mode == "add" and item.stack_group == operation.stack_group
                for key in item.row_keys
            }
            for row in range(group.start_row, body_end):
                label = str(worksheet.cell(row, operation.key_column or 1).value or "").strip()
                if label not in stacked_keys and operation.row_filter.matches(label):
                    leaked_filtered_rows.append(f"{operation.sheet}!{worksheet.cell(row, operation.key_column or 1).coordinate}")
    checks.append(
        {
            "name": "configured_row_filters",
            "passed": not leaked_filtered_rows,
            "cells": leaked_filtered_rows,
        }
    )

    marker_checks: list[dict[str, Any]] = []
    if spec:
        for group in spec.stack_groups:
            if group.retain_end_marker and group.end_marker_prefix:
                present = any(
                    str(workbook[group.sheet].cell(row, 1).value or "").startswith(group.end_marker_prefix)
                    for row in range(1, workbook[group.sheet].max_row + 1)
                )
                marker_checks.append({"group": group.id, "passed": present})
    checks.append({"name": "retained_markers", "passed": all(item["passed"] for item in marker_checks), "groups": marker_checks})
    workbook.close()
    return VerificationResult(passed=all(check["passed"] for check in checks), checks=checks)
