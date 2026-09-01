from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from tests.fixture_specs import PLAN_SHEET, STATUS_SHEET


HEADER_FILL = PatternFill("solid", fgColor="DCE6F1")


def _style_headers(worksheet, width: int) -> None:
    for row in worksheet.iter_rows(min_row=1, max_row=2, max_col=width):
        for cell in row:
            cell.font = Font(bold=True)
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center")
    worksheet.freeze_panes = "A3"
    worksheet.column_dimensions["A"].width = 38
    for column in range(2, width + 1):
        worksheet.column_dimensions[get_column_letter(column)].width = 12


def _status_sheet(workbook: Workbook, values: tuple[object, object, object, object],
                  *, template: bool = False) -> None:
    worksheet = workbook.active
    worksheet.title = STATUS_SHEET
    worksheet.append(["Category", "Current Data", "Current Data"])
    worksheet.append(["Category", "Metric 1", "Metric 2"])
    worksheet.append(["Category 1", values[0], values[1]])
    worksheet.append(["Category 2", values[2], values[3]])
    if template:
        worksheet["A8"] = (
            "Instruction: Match rows by category and add numeric values from all sources."
        )
    _style_headers(worksheet, 3)


def _plan_headers(worksheet, *, inserted_after_k: bool = False) -> dict[int, int]:
    target_to_source: dict[int, int] = {1: 1}
    worksheet.cell(1, 1).value = "Project Name"
    worksheet.cell(2, 1).value = "Project Name"
    for target_column in range(2, 26):
        source_column = target_column + (1 if inserted_after_k and target_column >= 12 else 0)
        target_to_source[target_column] = source_column
        worksheet.cell(1, source_column).value = "Plan Data"
        worksheet.cell(2, source_column).value = f"Field {target_column:02d}"
    if inserted_after_k:
        worksheet.cell(1, 12).value = "Source Extra Data"
        worksheet.cell(2, 12).value = "Inserted Field"
    return target_to_source


def _write_plan_row(worksheet, row: int, mapping: dict[int, int], label: str,
                    values: dict[int, object] | None = None) -> None:
    worksheet.cell(row, mapping[1]).value = label
    for target_column in range(2, 26):
        worksheet.cell(row, mapping[target_column]).value = (values or {}).get(target_column, 0)


def _plan_sheet(workbook: Workbook, project_name: str | None,
                *, template: bool = False, inserted_after_k: bool = False,
                project_values: dict[int, object] | None = None,
                aggregate_values: dict[str, dict[int, object]] | None = None) -> None:
    worksheet = workbook.create_sheet(PLAN_SHEET)
    mapping = _plan_headers(worksheet, inserted_after_k=inserted_after_k)
    if template:
        _write_plan_row(worksheet, 5, mapping, "Example: Project Name")
        _write_plan_row(worksheet, 6, mapping, "Replace with Project Name")
    elif project_name:
        _write_plan_row(worksheet, 5, mapping, project_name, project_values)
    aggregate_values = aggregate_values or {}
    _write_plan_row(
        worksheet, 7, mapping, "Project Type 1", aggregate_values.get("Project Type 1")
    )
    _write_plan_row(
        worksheet, 8, mapping, "Project Type 2", aggregate_values.get("Project Type 2")
    )
    worksheet.cell(9, 1).value = (
        "Instruction: Append project rows and add project-type totals by field."
    )
    _style_headers(worksheet, 26 if inserted_after_k else 25)


def _save_source(path: Path, status: tuple[object, object, object, object],
                 project_name: str, *, inserted_after_k: bool = False,
                 project_values: dict[int, object] | None = None,
                 aggregate_values: dict[str, dict[int, object]] | None = None) -> None:
    workbook = Workbook()
    _status_sheet(workbook, status)
    _plan_sheet(
        workbook, project_name, inserted_after_k=inserted_after_k,
        project_values=project_values, aggregate_values=aggregate_values,
    )
    workbook.save(path)
    workbook.close()


def build_synthetic_workbooks(directory: Path) -> Path:
    """Create small, fictional workbooks that exercise the production merge contract."""
    directory.mkdir(parents=True, exist_ok=True)

    template = Workbook()
    _status_sheet(template, (None, None, None, None), template=True)
    _plan_sheet(template, None, template=True)
    template.save(directory / "template.xlsx")
    template.close()

    _save_source(
        directory / "source_1.xlsx", (400, 100, 0, 0),
        "Project 1", project_values={16: 1327},
    )
    _save_source(
        directory / "source_2.xlsx", (26, 31, "/", "/"),
        "Project 2", project_values={16: 25},
    )
    _save_source(
        directory / "source_shifted.xlsx", (12, 8, 0, 0),
        "Project 3", inserted_after_k=True,
        project_values={12: 23, 16: 800, 18: 0, 25: 0},
        aggregate_values={"Project Type 1": {12: 6, 14: 7, 16: 88}},
    )
    return directory
