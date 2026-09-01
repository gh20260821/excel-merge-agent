from __future__ import annotations

from io import BytesIO
import json
from pathlib import Path
import sqlite3

import pytest
from openpyxl import Workbook, load_workbook
from starlette.datastructures import UploadFile

from app.domain import (
    ConflictResolution,
    HumanDecisionResponse,
    MergeOperation,
    MergeSpec,
    PlannerProvenance,
    RowFilterRules,
    RunRecord,
    RunState,
    StackGroup,
    VerificationResult,
)
from app.persistence import ConcurrentRunUpdate, RunRepository
from app.service import RunService
from app.workbook import build_planning_evidence, validate_merge_spec, verify_output
from tests.fixture_specs import PLAN_SHEET, STATUS_SHEET, representative_fixture_spec
from tests.synthetic_workbooks import build_synthetic_workbooks


@pytest.fixture(scope="session")
def representative_fixtures(tmp_path_factory: pytest.TempPathFactory) -> Path:
    return build_synthetic_workbooks(tmp_path_factory.mktemp("representative-workbooks"))


def upload(path: Path) -> UploadFile:
    return UploadFile(file=BytesIO(path.read_bytes()), filename=path.name)


def aggregate_row(worksheet: object, label: str) -> int:
    for row in range(1, worksheet.max_row + 1):
        if worksheet.cell(row, 1).value == label:
            return row
    raise AssertionError(f"Could not find aggregate row {label!r}")


def test_repository_rejects_stale_run_updates(tmp_path: Path) -> None:
    repository = RunRepository(tmp_path / "concurrency.sqlite3")
    current = repository.save(RunRecord(id="run-concurrency"))
    first = repository.get(current.id)
    stale = repository.get(current.id)
    first.add_event("first_update", "First writer wins.")
    repository.save(first)
    stale.add_event("stale_update", "This must not overwrite the first writer.")
    with pytest.raises(ConcurrentRunUpdate):
        repository.save(stale)


@pytest.mark.asyncio
async def test_representative_merge(tmp_path: Path, representative_fixtures: Path) -> None:
    service = RunService(RunRepository(tmp_path / "runs.sqlite3"), tmp_path / "runs")
    run = service.create()
    run = await service.save_files(
        run.id,
        upload(representative_fixtures / "template.xlsx"),
        [upload(representative_fixtures / "source_1.xlsx"), upload(representative_fixtures / "source_2.xlsx")],
    )
    assert run.state == RunState.FILES_UPLOADED

    run = service.inspect_and_plan(run.id, representative_fixture_spec())
    assert len(run.spec.operations) == 3
    assert len(run.conflicts) == 2
    assert {conflict.actual for conflict in run.conflicts} == {"/"}

    run = service.resolve_conflict(
        run.id,
        run.conflicts[0].id,
        ConflictResolution(action="treat_as_zero", scope="identical_in_run"),
    )
    assert all(conflict.resolved for conflict in run.conflicts)
    assert run.state == RunState.AWAITING_WRITE_APPROVAL
    run = service.approve(run.id, run.spec_hash or "")

    run = service.execute(run.id)
    assert run.state == RunState.COMPLETED
    assert run.verification and run.verification.passed
    assert run.compiled_plan and run.write_approval
    assert run.write_approval.compiled_plan_hash == run.compiled_plan.digest()
    check_names = {check["name"] for check in run.verification.checks}
    assert {"cell_reconciliation", "row_reconciliation", "untouched_template_regions"} <= check_names

    workbook = load_workbook(run.output_path, data_only=True)
    status = workbook[STATUS_SHEET]
    assert status["B3"].value == 426
    assert status["C3"].value == 131
    assert status["B4"].value == 0
    assert status["C4"].value == 0

    plan = workbook[PLAN_SHEET]
    assert plan["A5"].value == "Project 1"
    assert plan["A6"].value == "Project 2"
    assert plan["A7"].value == "Project Type 1"
    assert plan["A8"].value == "Project Type 2"
    assert str(plan["A9"].value).startswith("Instruction:")
    assert plan["P5"].value == 1327
    workbook.close()


@pytest.mark.asyncio
async def test_semantic_column_alignment_handles_inserted_source_column(
    tmp_path: Path, representative_fixtures: Path,
) -> None:
    service = RunService(RunRepository(tmp_path / "shifted.sqlite3"), tmp_path / "runs")
    run = service.create()
    run = await service.save_files(
        run.id,
        upload(representative_fixtures / "template.xlsx"),
        [upload(representative_fixtures / "source_shifted.xlsx")],
    )
    run = service.inspect_and_plan(run.id, representative_fixture_spec())
    assert run.conflicts == []
    compiled = next(
        item for item in run.compiled_plan.operations if item.operation_id == "independent_projects"
    )
    shifted_mapping = next(
        item for item in compiled.sources if item.source_file == "source_shifted.xlsx"
    )
    assert next(
        item.source_column for item in shifted_mapping.columns if item.target_column == 12
    ) == 13
    run = service.approve(run.id, run.spec_hash or "")
    run = service.execute(run.id)
    assert run.state == RunState.COMPLETED

    workbook = load_workbook(run.output_path, data_only=True)
    sheet = workbook[PLAN_SHEET]
    # The synthetic shifted source inserts one field after K. Header-path
    # alignment maps source M/Q/S/Z back to template L/P/R/Y.
    assert sheet["A5"].value == "Project 3"
    assert sheet["L5"].value == 23
    assert sheet["P5"].value == 800
    assert sheet["R5"].value == 0
    assert sheet["Y5"].value == 0
    aggregate_row_number = aggregate_row(sheet, "Project Type 1")
    assert sheet.cell(aggregate_row_number, 12).value == 6
    assert sheet.cell(aggregate_row_number, 14).value == 7
    assert sheet.cell(aggregate_row_number, 16).value == 88
    workbook.close()


@pytest.mark.asyncio
async def test_duplicate_source_content_is_blocked_before_write(
    tmp_path: Path, representative_fixtures: Path
) -> None:
    service = RunService(RunRepository(tmp_path / "duplicates.sqlite3"), tmp_path / "runs")
    run = service.create()
    run = await service.save_files(
        run.id,
        upload(representative_fixtures / "template.xlsx"),
        [upload(representative_fixtures / "source_1.xlsx"), upload(representative_fixtures / "source_1.xlsx")],
    )
    run = service.inspect_and_plan(run.id, representative_fixture_spec())
    duplicate = next(item for item in run.conflicts if item.type == "duplicate_source")
    assert duplicate.source_id
    assert run.state == RunState.AWAITING_USER_INPUT


@pytest.mark.asyncio
async def test_duplicate_business_key_is_never_resolved_by_first_match(
    tmp_path: Path, representative_fixtures: Path
) -> None:
    source_path = tmp_path / "duplicate-key.xlsx"
    workbook = load_workbook(representative_fixtures / "source_shifted.xlsx")
    sheet = workbook[STATUS_SHEET]
    sheet["A5"] = "Category 1"
    workbook.save(source_path)
    workbook.close()
    service = RunService(RunRepository(tmp_path / "keys.sqlite3"), tmp_path / "runs")
    run = service.create()
    run = await service.save_files(
        run.id,
        upload(representative_fixtures / "template.xlsx"),
        [upload(source_path)],
    )
    run = service.inspect_and_plan(run.id, representative_fixture_spec())
    conflict = next(item for item in run.conflicts if item.type == "duplicate_row_key")
    assert "duplicated" in conflict.message
    assert run.compiled_plan
    assert run.state == RunState.AWAITING_USER_INPUT


@pytest.mark.asyncio
async def test_reconciliation_detects_tampered_output(
    tmp_path: Path, representative_fixtures: Path
) -> None:
    service = RunService(RunRepository(tmp_path / "tamper.sqlite3"), tmp_path / "runs")
    run = service.create()
    run = await service.save_files(
        run.id,
        upload(representative_fixtures / "template.xlsx"),
        [upload(representative_fixtures / "source_shifted.xlsx")],
    )
    run = service.inspect_and_plan(run.id, representative_fixture_spec())
    run = service.approve(run.id, run.spec_hash or "")
    run = service.execute(run.id)
    manifest = json.loads(Path(run.audit_path).read_text(encoding="utf-8"))["reconciliation"]
    workbook = load_workbook(run.output_path)
    workbook[PLAN_SHEET]["L5"] = 999999
    workbook.save(run.output_path)
    workbook.close()
    verification = verify_output(
        Path(run.output_path),
        run.spec,
        manifest=manifest,
        template_path=Path(run.template.stored_path),
    )
    assert not verification.passed
    cell_check = next(
        item for item in verification.checks if item["name"] == "row_reconciliation"
    )
    assert not cell_check["passed"]


@pytest.mark.asyncio
async def test_aggregate_text_is_blocking_and_can_be_resolved(
    tmp_path: Path, representative_fixtures: Path
) -> None:
    source_path = tmp_path / "source-with-aggregate-marker.xlsx"
    source_book = load_workbook(representative_fixtures / "source_1.xlsx")
    source_sheet = source_book[PLAN_SHEET]
    marker_row = aggregate_row(source_sheet, "Project Type 1")
    source_sheet.cell(marker_row, 2).value = "/"
    source_book.save(source_path)
    source_book.close()

    service = RunService(RunRepository(tmp_path / "runs.sqlite3"), tmp_path / "runs")
    run = service.create()
    run = await service.save_files(
        run.id,
        upload(representative_fixtures / "template.xlsx"),
        [upload(source_path)],
    )
    run = service.inspect_and_plan(run.id, representative_fixture_spec())

    aggregate_conflicts = [
        conflict
        for conflict in run.conflicts
        if conflict.sheet == PLAN_SHEET
    ]
    assert len(aggregate_conflicts) == 1
    assert aggregate_conflicts[0].row_key == "Project Type 1"
    assert aggregate_conflicts[0].actual == "/"
    assert aggregate_conflicts[0].recommended_action == "treat_as_zero"

    assert run.state == RunState.AWAITING_USER_INPUT
    with pytest.raises(ValueError, match="Answer 1 question"):
        service.approve(run.id, run.spec_hash or "")

    run = service.resolve_conflict(
        run.id,
        aggregate_conflicts[0].id,
        ConflictResolution(action="treat_as_zero", scope="this_cell"),
    )
    assert run.state == RunState.AWAITING_WRITE_APPROVAL
    run = service.approve(run.id, run.spec_hash or "")
    run = service.execute(run.id)
    assert run.state == RunState.COMPLETED

    output_book = load_workbook(run.output_path, data_only=True)
    output_sheet = output_book[PLAN_SHEET]
    output_row = aggregate_row(output_sheet, "Project Type 1")
    assert output_sheet.cell(output_row, 2).value == 0
    output_book.close()


def test_validator_rejects_add_columns_outside_template(representative_fixtures: Path) -> None:
    spec = representative_fixture_spec()
    status_operation = next(item for item in spec.operations if item.id == "status_add")
    status_operation.value_columns = [999]
    with pytest.raises(ValueError, match="outside the template"):
        validate_merge_spec(spec, representative_fixtures / "template.xlsx")


def test_planning_evidence_contains_template_guidelines(representative_fixtures: Path) -> None:
    evidence = build_planning_evidence(
        representative_fixtures / "template.xlsx",
        [representative_fixtures / "source_1.xlsx", representative_fixtures / "source_2.xlsx"],
    )
    template_text = " ".join(
        cell["value"]
        for sheet in evidence["template_sheets"]
        for cell in sheet["nonempty_cells"]
    )
    assert "Instruction:" in template_text
    assert evidence["executor_contract"]["configuration_driven"] is True
    assert evidence["executor_contract"]["supported_modes"] == ["add", "concatenate"]
    assert evidence["source_population"]["total_workbooks"] == 2
    assert evidence["source_population"]["representatives_in_prompt"] <= 2
    assert sum(group["count"] for group in evidence["source_population"]["schema_groups"]) == 2


def test_plan_validator_rejects_aggregate_rows_in_concatenate_region(
    representative_fixtures: Path,
) -> None:
    spec = representative_fixture_spec()
    concatenate = next(item for item in spec.operations if item.id == "independent_projects")
    concatenate.row_filter.exclude_exact_values = []
    with pytest.raises(ValueError, match="prevent duplicates"):
        validate_merge_spec(spec, representative_fixtures / "template.xlsx")


def test_plan_validator_rejects_duplicate_concatenate_coverage(
    representative_fixtures: Path,
) -> None:
    spec = representative_fixture_spec()
    concatenate = next(item for item in spec.operations if item.mode == "concatenate")
    duplicate = concatenate.model_copy(deep=True)
    duplicate.id = "duplicate_source_copy"
    duplicate.stack_order = 11
    spec.operations.append(duplicate)
    with pytest.raises(ValueError, match="copy the same rows"):
        validate_merge_spec(spec, representative_fixtures / "template.xlsx")


def test_plan_validator_requires_template_body_classification(
    representative_fixtures: Path,
) -> None:
    spec = representative_fixture_spec()
    concatenate = next(item for item in spec.operations if item.mode == "concatenate")
    concatenate.row_filter = RowFilterRules(
        exclude_exact_values=["Project Type 1", "Project Type 2"]
    )
    with pytest.raises(ValueError, match="template body rows unclassified"):
        validate_merge_spec(spec, representative_fixtures / "template.xlsx")


def test_plan_validator_rejects_invalid_filter_regex(representative_fixtures: Path) -> None:
    spec = representative_fixture_spec()
    concatenate = next(item for item in spec.operations if item.mode == "concatenate")
    concatenate.row_filter.exclude_regexes = ["["]
    with pytest.raises(ValueError, match="invalid row-filter regex"):
        validate_merge_spec(spec, representative_fixtures / "template.xlsx")


def test_plan_validator_rejects_early_concatenate_boundary(
    representative_fixtures: Path,
) -> None:
    spec = representative_fixture_spec()
    concatenate = next(item for item in spec.operations if item.mode == "concatenate")
    concatenate.end_marker_prefix = "Example: Project"
    with pytest.raises(ValueError, match="ends before the full stack-group body"):
        validate_merge_spec(spec, representative_fixtures / "template.xlsx")


def test_conversation_snapshot_persists_across_service_restart(tmp_path: Path) -> None:
    database_path = tmp_path / "conversation.sqlite3"
    run_root = tmp_path / "conversation-runs"
    service = RunService(RunRepository(database_path), run_root)
    run = service.create()
    messages = [
        {
            "id": "message-1",
            "role": "user",
            "metadata": {"createdAt": "2026-08-25T00:00:00Z"},
            "parts": [{"type": "text", "text": "Merge these workbooks."}],
        }
    ]
    service.save_conversation(run.id, messages)

    restarted = RunService(RunRepository(database_path), run_root)
    assert restarted.get(run.id).conversation == messages


def test_run_list_skips_an_incompatible_legacy_record(tmp_path: Path) -> None:
    database_path = tmp_path / "legacy-list.sqlite3"
    repository = RunRepository(database_path)
    repository.save(RunRecord(id="valid-run"))
    with sqlite3.connect(database_path) as connection:
        connection.execute(
            "INSERT INTO runs(id, state, updated_at, payload) VALUES (?, ?, ?, ?)",
            ("legacy-run", "created", "2026-08-25T00:00:00Z", '{"id":"legacy-run","state":"unknown"}'),
        )

    assert [run.id for run in repository.list()] == ["valid-run"]


def test_legacy_resolved_human_state_migrates_to_single_write_gate(tmp_path: Path) -> None:
    repository = RunRepository(tmp_path / "legacy-state.sqlite3")
    run = RunRecord(id="legacy-resolved", state=RunState.AWAITING_HUMAN)
    repository.save(run)

    service = RunService(repository, tmp_path / "legacy-runs")
    migrated = service.get(run.id)

    assert migrated.state == RunState.AWAITING_WRITE_APPROVAL
    assert migrated.approved_spec_hash is None
    assert migrated.write_approval is None
    assert migrated.events[-1]["kind"] == "legacy_state_migrated"


def test_unfinished_plan_hash_upgrades_for_new_alignment_default(tmp_path: Path) -> None:
    repository = RunRepository(tmp_path / "schema-upgrade.sqlite3")
    spec = generic_merge_spec()
    run = RunRecord(
        id="old-plan-hash",
        state=RunState.AWAITING_WRITE_APPROVAL,
        spec=spec,
        spec_hash="legacy-digest",
        approved_spec_hash="legacy-digest",
    )
    repository.save(run)

    migrated = RunService(repository, tmp_path / "schema-runs").get(run.id)

    assert migrated.spec_hash == spec.digest()
    assert migrated.approved_spec_hash is None
    assert migrated.state == RunState.AWAITING_WRITE_APPROVAL
    assert migrated.events[-1]["kind"] == "plan_schema_upgraded"


def create_generic_workbooks(tmp_path: Path) -> tuple[Path, list[Path]]:
    template_path = tmp_path / "template.xlsx"
    template = Workbook()
    inventory = template.active
    inventory.title = "Inventory"
    inventory.append(["Item", "Units", "Cost"])
    inventory.append(["Apples", None, None])
    inventory.append(["Pears", None, None])
    inventory["A5"] = "Guideline: add inventory; concatenate project rows."
    projects = template.create_sheet("Projects")
    projects.append(["Project", "Hours"])
    projects.append(["replace with project", None])
    projects.append(["NOTE:", "Retain this note"])
    template.save(template_path)
    template.close()

    source_paths: list[Path] = []
    for name, inventory_rows, project_row in [
        ("north.xlsx", [("Apples", 2, 5), ("Pears", 3, 7)], ("North rollout", 10)),
        ("south.xlsx", [("Apples", 4, 11), ("Pears", 1, 13)], ("South rollout", 20)),
    ]:
        path = tmp_path / name
        source = Workbook()
        sheet = source.active
        sheet.title = "Inventory"
        sheet.append(["Item", "Units", "Cost"])
        for row in inventory_rows:
            sheet.append(row)
        project = source.create_sheet("Projects")
        project.append(["Project", "Hours"])
        project.append(project_row)
        project.append(["NOTE:", "Retain this note"])
        source.save(path)
        source.close()
        source_paths.append(path)
    return template_path, source_paths


def generic_merge_spec() -> MergeSpec:
    return MergeSpec(
        template_family="inventory_demo_v1",
        operations=[
            MergeOperation(
                id="inventory_totals",
                sheet="Inventory",
                mode="add",
                description="Add inventory totals by item.",
                row_keys=["Apples", "Pears"],
                alignment="row_key",
                key_column=1,
                value_columns=[2, 3],
            ),
            MergeOperation(
                id="project_rows",
                sheet="Projects",
                mode="concatenate",
                description="Concatenate source projects.",
                alignment="position",
                key_column=1,
                data_start_row=2,
                end_marker_prefix="NOTE:",
                column_count=2,
                row_filter=RowFilterRules(
                    exclude_exact_values=["replace with project"]
                ),
                style_template_row=2,
                placement="stack",
                stack_group="projects_body",
                stack_order=10,
            ),
        ],
        stack_groups=[
            StackGroup(
                id="projects_body",
                sheet="Projects",
                start_row=2,
                column_count=2,
                end_marker_prefix="NOTE:",
            )
        ],
        rationale="The template guideline defines two independent operations.",
        guideline_citations=["Inventory!A5"],
    )


@pytest.mark.asyncio
async def test_model_planner_retries_an_empty_tool_submission(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    from app import planner as planner_module

    template_path, _ = create_generic_workbooks(tmp_path)
    expected = generic_merge_spec()

    class FakePlanningStream:
        def __init__(self, call_number: int) -> None:
            self.call_number = call_number

        async def __aenter__(self):
            return self

        async def __aexit__(self, *_: object) -> None:
            return None

        def __aiter__(self):
            async def events():
                if self.call_number == 2:
                    context = planner_module._planning_context.get()
                    assert context is not None
                    context.proposals.append(expected)
                if False:
                    yield None

            return events()

    class FakePlanningAgent:
        calls = 0

        def run(self, *_: object, **__: object) -> FakePlanningStream:
            self.calls += 1
            return FakePlanningStream(self.calls)

    fake_agent = FakePlanningAgent()
    monkeypatch.setattr(planner_module, "planning_agent", fake_agent)
    planned, provenance = await planner_module.create_model_plan(
        object(), "test-model", template_path, {"evidence": "test"}
    )

    assert planned == expected
    assert provenance.attempts == 2
    assert fake_agent.calls == 2


@pytest.mark.asyncio
async def test_configuration_driven_executor_supports_other_layout(tmp_path: Path) -> None:
    template_path, source_paths = create_generic_workbooks(tmp_path)
    spec = generic_merge_spec()

    service = RunService(RunRepository(tmp_path / "generic.sqlite3"), tmp_path / "runs")
    run = service.create()
    run = await service.save_files(
        run.id,
        upload(template_path),
        [upload(path) for path in source_paths],
    )
    service.prepare_inspection(run.id)
    run = service.accept_plan(
        run.id,
        spec,
        PlannerProvenance(kind="test", model="generic-test", evidence_sha256="test"),
    )
    assert run.conflicts == []
    run = service.approve(run.id, run.spec_hash or "")
    run = service.execute(run.id, supplied_spec=spec)

    workbook = load_workbook(run.output_path, data_only=True)
    assert workbook["Inventory"]["B2"].value == 6
    assert workbook["Inventory"]["C2"].value == 16
    assert workbook["Inventory"]["B3"].value == 4
    assert workbook["Projects"]["A2"].value == "North rollout"
    assert workbook["Projects"]["A3"].value == "South rollout"
    assert workbook["Projects"]["A4"].value == "NOTE:"
    workbook.close()

    drifted = spec.model_copy(deep=True)
    drifted.operations[0].value_columns = [2]
    with pytest.raises(ValueError, match="does not match the approved plan"):
        service.execute(run.id, supplied_spec=drifted)


@pytest.mark.asyncio
async def test_batch_execution_preserves_results_and_persists_progress(
    tmp_path: Path,
) -> None:
    template_path, source_paths = create_generic_workbooks(tmp_path)
    service = RunService(RunRepository(tmp_path / "batches.sqlite3"), tmp_path / "runs")
    run = service.create()
    run = service.configure_batch(run.id, 1)
    assert run.batch_size == 1
    run = await service.save_files(
        run.id,
        upload(template_path),
        [upload(path) for path in source_paths],
    )
    service.prepare_inspection(run.id)
    run = service.accept_plan(
        run.id,
        generic_merge_spec(),
        PlannerProvenance(kind="test", model="batch-test", evidence_sha256="test"),
    )
    run = service.approve(run.id, run.spec_hash or "")
    assert run.write_approval and run.write_approval.batch_size == 1
    run = service.execute(run.id)

    assert run.state == RunState.COMPLETED
    assert run.batch_progress
    assert run.batch_progress.status == "completed"
    assert run.batch_progress.total_sources == 2
    assert run.batch_progress.total_work_units == 4
    assert run.batch_progress.completed_work_units == 4
    assert run.batch_progress.processed_sources == 2
    assert {event["kind"] for event in run.events} >= {
        "batch_configuration_updated",
        "batch_execution_started",
        "batch_execution_completed",
    }
    audit = json.loads(Path(run.audit_path).read_text(encoding="utf-8"))
    assert audit["batch_execution"] == {
        "batch_size": 1,
        "active_source_count": 2,
    }

    workbook = load_workbook(run.output_path, data_only=True)
    assert workbook["Inventory"]["B2"].value == 6
    assert workbook["Projects"]["A2"].value == "North rollout"
    assert workbook["Projects"]["A3"].value == "South rollout"
    workbook.close()


def test_batch_change_invalidates_existing_write_approval(tmp_path: Path) -> None:
    service = RunService(RunRepository(tmp_path / "batch-approval.sqlite3"), tmp_path / "runs")
    run = service.create()
    run.batch_size = 10
    run.state = RunState.PLAN_READY
    # A minimal persisted grant is enough to exercise configuration invalidation.
    from app.domain import WriteApprovalGrant

    run.approved_spec_hash = "spec"
    run.spec_hash = "spec"
    run.write_approval = WriteApprovalGrant(
        spec_hash="spec",
        template_sha256="template",
        source_sha256s={},
        output_paths=["merged.xlsx", "audit.json"],
        batch_size=10,
    )
    service.repository.save(run)

    changed = service.configure_batch(run.id, 25)
    assert changed.batch_size == 25
    assert changed.write_approval is None
    assert changed.approved_spec_hash is None
    assert changed.state == RunState.AWAITING_WRITE_APPROVAL


@pytest.mark.asyncio
async def test_changed_input_invalidates_write_approval_without_asking(tmp_path: Path) -> None:
    template_path, source_paths = create_generic_workbooks(tmp_path)
    spec = generic_merge_spec()
    database_path = tmp_path / "resume.sqlite3"
    run_root = tmp_path / "resume-runs"
    service = RunService(RunRepository(database_path), run_root)
    run = service.create()
    run = await service.save_files(
        run.id,
        upload(template_path),
        [upload(path) for path in source_paths],
    )
    service.prepare_inspection(run.id)
    run = service.accept_plan(
        run.id,
        spec,
        PlannerProvenance(kind="test", model="resume-test", evidence_sha256="test"),
    )
    run = service.approve(run.id, run.spec_hash or "")

    affected_source = next(item for item in run.sources if item.filename == "north.xlsx")
    changed = load_workbook(affected_source.stored_path)
    changed.remove(changed["Projects"])
    changed.save(affected_source.stored_path)
    changed.close()

    replanning = service.execute(run.id, supplied_spec=spec)
    assert replanning.state == RunState.AWAITING_USER_INPUT
    assert replanning.execution_attempts == 1
    assert replanning.output_path is None
    assert replanning.write_approval is None
    assert replanning.approved_spec_hash is None
    assert replanning.decisions == []
    assert any(conflict.type == "missing_sheet" for conflict in replanning.conflicts)
    assert not (run_root / run.id / "outputs" / "merged.xlsx").exists()
    assert replanning.events[-1]["kind"] == "input_change_detected"

    restarted_service = RunService(RunRepository(database_path), run_root)
    persisted = restarted_service.get(run.id)
    assert persisted.state == RunState.AWAITING_USER_INPUT
    assert persisted.decisions == []


@pytest.mark.asyncio
async def test_verification_failure_rebuilds_before_failing(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    template_path, source_paths = create_generic_workbooks(tmp_path)
    spec = generic_merge_spec()
    service = RunService(RunRepository(tmp_path / "verify.sqlite3"), tmp_path / "verify-runs")
    run = service.create()
    run = await service.save_files(
        run.id,
        upload(template_path),
        [upload(path) for path in source_paths],
    )
    service.prepare_inspection(run.id)
    run = service.accept_plan(
        run.id,
        spec,
        PlannerProvenance(kind="test", model="verify-test", evidence_sha256="test"),
    )
    run = service.approve(run.id, run.spec_hash or "")

    monkeypatch.setattr(
        "app.service.execute_merge",
        lambda *_args, **_kwargs: VerificationResult(
            passed=False,
            checks=[{"name": "injected_check", "passed": False}],
        ),
    )
    failed = service.execute(run.id, supplied_spec=spec)
    assert failed.state == RunState.FAILED
    assert failed.decisions == []
    assert len(failed.recovery_attempts) == 1
    assert failed.recovery_attempts[0].action == "rebuild_from_untouched_template"
    assert failed.recovery_attempts[0].outcome == "failed"
    assert not (tmp_path / "verify-runs" / run.id / "outputs" / "merged.xlsx").exists()


@pytest.mark.asyncio
async def test_transient_executor_error_retries_under_same_write_grant(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    from app import service as service_module

    template_path, source_paths = create_generic_workbooks(tmp_path)
    spec = generic_merge_spec()
    service = RunService(RunRepository(tmp_path / "retry.sqlite3"), tmp_path / "retry-runs")
    run = service.create()
    run = await service.save_files(
        run.id, upload(template_path), [upload(path) for path in source_paths]
    )
    service.prepare_inspection(run.id)
    run = service.accept_plan(
        run.id,
        spec,
        PlannerProvenance(kind="test", model="retry-test", evidence_sha256="test"),
    )
    run = service.approve(run.id, run.spec_hash or "")
    grant_time = run.write_approval.granted_at

    real_execute_merge = service_module.execute_merge
    calls = 0

    def flaky_execute(*args, **kwargs):
        nonlocal calls
        calls += 1
        if calls == 1:
            raise OSError("temporary read failure")
        return real_execute_merge(*args, **kwargs)

    monkeypatch.setattr(service_module, "execute_merge", flaky_execute)
    completed = service.execute(run.id, supplied_spec=spec)

    assert completed.state == RunState.COMPLETED
    assert calls == 2
    assert completed.write_approval.granted_at == grant_time
    assert len(completed.recovery_attempts) == 1
    assert completed.recovery_attempts[0].category == "transient"
    assert completed.recovery_attempts[0].outcome == "succeeded"
